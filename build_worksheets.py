"""
    Build worksheets from input data ... 
    - input data has transactions sorted by date
    - read the input and creat one file for each date and put all transactions into the file
    - output filename = DPYYMMDD.xlsx

    logic
    open input_data.xlsx

    old_date = ""
    read input_data sheet
    for each row in input_data
        if date <> old_date
            if old_date is not null 
                save DPYYMMDD.xlsx
            create new file DPYYMMDD.xlsx
        append row to output_file
    
    save DPYYMMDD.xlsx

"""

import openpyxl
import datetime



def main():
    input_file = 'input_data.xlsx'
    output_file_template = 'FUPLOAD_Template.xlsx'

    # Load the workbook and select the active sheet
    wb: openpyxl.Workbook = openpyxl.load_workbook(input_file)
    ws = wb['input_data']

    old_date = datetime.datetime(1980, 1, 1)
    current_wb = None
    current_ws = None


    for row in ws.iter_rows(min_row=2, values_only=True):
        date = row[0]  # Assuming the first column is the date
        print(f"Date: {date}   Amount: {row[1]}")
        if old_date == datetime.datetime(1980, 1, 1) or date != old_date:
            if current_wb is not None:
                fname:str = "DP" + old_date.strftime('%y%m%d') + ".xlsx"
                current_wb.save(fname)
                current_wb.close()

            old_date = date
            cur_row = 19
            current_wb = openpyxl.load_workbook(output_file_template)
            current_ws = current_wb["Sheet"]
            # current_ws.append(ws[1])  # Append header row

        current_ws.cell(cur_row, 2).value = row[3]
        current_ws.cell(cur_row, 3).value = row[4]
        current_ws.cell(cur_row, 8).value = row[1]
        current_ws.cell(cur_row,10).value = row[2]
        cur_row += 1

    if current_wb is not None:
        fname:str = "DP" + old_date.strftime('%y%m%d') + ".xlsx"
        current_wb.save(fname)
        current_wb.close()



if __name__ == "__main__":
    # main("debug")
    # main("prod")
    main()



